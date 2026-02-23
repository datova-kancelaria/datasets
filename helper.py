from __future__ import annotations

import json
import random
import sys
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Set

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ----------------------------
# defaults
# ----------------------------

_HEADERS_DEFAULT: dict[str, str] = {
    "Accept": "application/json",
    "Content-Type": "application/json",
    "User-Agent": "oCDK_fetcher_agent",
}

_HEADERS_GET_DEFAULT: dict[str, str] = {
    "Accept": "application/json",
    "User-Agent": "oCDK_fetcher_agent",
}

_SESSION = requests.Session()

_RETRY_STATUS = {429, 500, 502, 503, 504}
_RETRY_EXC = (
    requests.exceptions.ProxyError,
    requests.exceptions.ConnectionError,
    requests.exceptions.Timeout,
)

# ----------------------------
# HTTP helpers (GET/POST + retry)
# ----------------------------

def request_with_retry(
    method: str,
    url: str,
    *,
    max_attempts: int = 6,
    base_sleep: float = 1.0,
    max_sleep: float = 15.0,
    **kwargs,
) -> requests.Response:
    """
    Retries on transient network/proxy errors and on selected HTTP status codes.
    """
    for attempt in range(1, max_attempts + 1):
        try:
            resp = _SESSION.request(method, url, **kwargs)

            if resp.status_code in _RETRY_STATUS and attempt < max_attempts:
                # free the connection before sleeping
                resp.close()

                ra = resp.headers.get("Retry-After")
                if ra is not None:
                    try:
                        sleep_s = float(ra)
                    except ValueError:
                        sleep_s = base_sleep
                else:
                    sleep_s = base_sleep * (2 ** (attempt - 1))

                sleep_s = min(max_sleep, sleep_s) * random.uniform(0.8, 1.2)
                print(
                    f"[retry] {method} {url} -> HTTP {resp.status_code}; sleeping {sleep_s:.1f}s "
                    f"(attempt {attempt}/{max_attempts})",
                    file=sys.stderr,
                    flush=True,
                )
                time.sleep(sleep_s)
                continue

            return resp

        except _RETRY_EXC as e:
            if attempt >= max_attempts:
                raise
            sleep_s = min(max_sleep, base_sleep * (2 ** (attempt - 1))) * random.uniform(0.8, 1.2)
            print(
                f"[retry] {method} {url} failed: {type(e).__name__}: {e}; sleeping {sleep_s:.1f}s "
                f"(attempt {attempt}/{max_attempts})",
                file=sys.stderr,
                flush=True,
            )
            time.sleep(sleep_s)

    raise RuntimeError("unreachable")


def get_json(
    url: str,
    *,
    headers: dict[str, str] | None = None,
    timeout: tuple[float, float] = (30.0, 60.0),
    verify_tls: bool = True,
) -> Any:
    hdrs = dict(_HEADERS_GET_DEFAULT)
    if headers:
        hdrs.update(headers)

    r = request_with_retry(
        "GET",
        url,
        headers=hdrs,
        timeout=timeout,
        allow_redirects=True,
        verify=verify_tls,
    )
    r.raise_for_status()
    try:
        return r.json()
    except ValueError as e:
        raise RuntimeError(
            "Response was not valid JSON. "
            f"Content-Type={r.headers.get('Content-Type')!r}. "
            f"First 200 chars:\n{(r.text or '')[:200]}"
        ) from e


def post_json(
    url: str,
    payload: dict[str, Any],
    *,
    headers: dict[str, str] | None = None,
    timeout: tuple[float, float] = (60.0, 60.0),
    verify_tls: bool = True,
) -> Any:
    hdrs = dict(_HEADERS_DEFAULT)
    if headers:
        hdrs.update(headers)

    r = request_with_retry(
        "POST",
        url,
        json=payload,
        headers=hdrs,
        timeout=timeout,
        allow_redirects=False,
        verify=verify_tls,
    )
    r.raise_for_status()
    try:
        return r.json()
    except ValueError as e:
        raise RuntimeError(
            "Response was not valid JSON. "
            f"Content-Type={r.headers.get('Content-Type')!r}. "
            f"First 200 chars:\n{(r.text or '')[:200]}"
        ) from e


# ----------------------------
# CMDB paging
# ----------------------------

def fetch_node(
    api_url: str,
    entity_name: str,
    *,
    limit: int | None = None,
    offset: int | None = None,
) -> Any:
    payload: Dict[str, Any] = {
        "parameters": {"target": "nodes", "type": entity_name, "validOnly": True}
    }
    if limit is not None:
        payload["parameters"]["limit"] = limit
    if offset is not None:
        payload["parameters"]["offset"] = offset
    return post_json(api_url, payload)


def fetch_paged(
    api_url: str,
    entity_name: str,
    *,
    page_size: int = 1000,
    sleep_ms: int = 1000,
    verbose: bool = True,
) -> List[Dict[str, Any]]:
    offset = 0
    out: List[Dict[str, Any]] = []
    page = 0

    while True:
        page += 1
        t0 = time.perf_counter()

        data = fetch_node(api_url, entity_name, limit=page_size, offset=offset)
        partial = data.get("result") or []
        dt = time.perf_counter() - t0

        if verbose:
            print(
                f"[{entity_name}] page={page} offset={offset} got={len(partial)} total={len(out) + len(partial)} "
                f"({dt:.2f}s)",
                flush=True,
            )

        if not partial:
            break

        out.extend(partial)
        offset += len(partial)

        if len(partial) < page_size:
            break

        if sleep_ms:
            time.sleep(sleep_ms / 1000.0)

    return out


# ----------------------------
# MetaIS metadata + enums
# ----------------------------

def normalize_enum(enum_raw: Dict[str, Any]) -> Dict[str, str]:
    items = enum_raw.get("enumItems") or []
    res: Dict[str, str] = {}
    for item in items:
        key = item.get("code")
        value = item.get("value")
        if key is not None:
            res[str(key)] = "" if value is None else str(value)
    return res


def fetch_enum(enum_name: str) -> Dict[str, str]:
    url = "https://metais.slovensko.sk/api/enums-repo/enums/enum/all/" + enum_name
    return normalize_enum(get_json(url))


def prepare_payload_attr(
    attr: Dict[str, Any],
    *,
    valid_override: bool | None = None,
) -> tuple[str | None, Dict[str, Any] | None]:
    key = attr.get("technicalName")
    name = attr.get("name")
    desc = attr.get("description")
    valid = attr.get("valid")

    if not key:
        return None, None

    if valid_override is not None:
        valid = valid_override

    payload: Dict[str, Any] = {
        "name": name,
        "description": desc,
        "valid": valid,
    }

    cons = attr.get("constraints", []) or []
    for constraint in cons:
        if constraint.get("type") == "enum":
            enum_code = constraint.get("enumCode")
            if enum_code:
                payload["enum"] = enum_code
                break

    return str(key), payload


def normalize_attr_metadata(metadata_raw: Dict[str, Any]) -> Dict[str, Any]:
    res: Dict[str, Any] = {}

    for attr in metadata_raw.get("attributes", []) or []:
        key, attr_sanitized = prepare_payload_attr(attr)
        if key and attr_sanitized is not None:
            res[key] = attr_sanitized

    for attrProfile in metadata_raw.get("attributeProfiles", []) or []:
        valid = attrProfile.get("valid", True)
        valid_override = False if (valid is False) else None
        for attr in attrProfile.get("attributes", []) or []:
            key, attr_sanitized = prepare_payload_attr(attr, valid_override=valid_override)
            if key and attr_sanitized is not None:
                res[key] = attr_sanitized

    return res


def fetch_attr_metadata(entity_name: str) -> Dict[str, Any]:
    url = "https://metais.slovensko.sk/api/types-repo/citypes/citype/" + entity_name
    return normalize_attr_metadata(get_json(url))


def apply_enum_recursive(value: Any, enum_map: Dict[str, str]) -> Tuple[Any, List[str]]:
    missing: List[str] = []

    def rec(x: Any) -> Any:
        if x is None:
            return None

        if isinstance(x, str):
            if x in enum_map:
                return enum_map[x]
            missing.append(x)
            return x

        if isinstance(x, (list, tuple)):
            return [rec(v) for v in x]

        if isinstance(x, dict):
            if "code" in x and isinstance(x["code"], str):
                code = x["code"]
                if code in enum_map:
                    y = dict(x)
                    y["value"] = enum_map[code]
                    return y
                missing.append(code)
            return x

        return x

    mapped = rec(value)
    return mapped, missing


def normalize_attributes(
    attributes: List[Dict[str, Any]],
    metadata: Dict[str, Any],
    enums: Dict[str, Dict[str, str]],
) -> Dict[str, Any]:
    res: Dict[str, Any] = {}

    for attr in attributes:
        attr_name = attr.get("name")
        if not attr_name:
            continue

        if "value" not in attr:
            continue
        attr_value = attr.get("value")

        attr_meta = metadata.get(attr_name)
        if not attr_meta:
            continue

        if attr_meta.get("valid") is not True:
            continue

        enum_name = attr_meta.get("enum")
        if enum_name:
            enum_map = enums.get(enum_name)
            if enum_map:
                attr_value, missing = apply_enum_recursive(attr_value, enum_map)
                if missing:
                    print(
                        f"Warning: enum {enum_name} for attribute {attr_name} missing keys: {sorted(set(missing))}",
                        flush=True,
                    )
            else:
                print(f"Warning: enum {enum_name} for attribute {attr_name} not loaded", flush=True)

        res[attr_name] = attr_value

    return res


def accept_entity_cloud_service(entity: Dict[str, Any]) -> bool:
    entity_type = entity.get("type", "")
    cloud_key = "EA_Profil_" + entity_type + "_typ_cloudovej_sluzby"
    cloud_type = (entity.get("attributes") or {}).get(cloud_key)
    return bool(cloud_type and cloud_type != "Žiadny")


def sanitize_node(
    data: List[Dict[str, Any]],
    metadata: Dict[str, Any],
    enums: Dict[str, Dict[str, str]],
) -> List[Dict[str, Any]]:
    sanitized: List[Dict[str, Any]] = []

    for entity in data:
        meta = entity.get("metaAttributes") or {}
        if meta.get("state") == "INVALIDATED":
            continue

        e = dict(entity)
        e.pop("metaAttributes", None)

        e["attributes"] = normalize_attributes(e.get("attributes", []) or [], metadata, enums)

        if not accept_entity_cloud_service(e):
            continue

        sanitized.append(e)

    return sanitized


def get_enums_needed(
    metadata: Dict[str, Any],
    enums: Dict[str, Dict[str, str]],
    *,
    sleep_s: float = 0.1,
) -> None:
    for _, attr_properties in metadata.items():
        enum_name = attr_properties.get("enum")
        if enum_name and enum_name not in enums:
            enums[enum_name] = fetch_enum(enum_name)
            time.sleep(sleep_s)


# ----------------------------
# Schema load + metadata merge + remap
# ----------------------------

Meta = Dict[str, Any]
MetaMap = Dict[str, Meta]
AttrMap = Dict[str, Any]
SchemaItem = Dict[str, str]

def load_merge_schema(path: Path) -> List[SchemaItem]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list):
        raise ValueError(f"Schema {path} must be a JSON list")
    out: List[SchemaItem] = []
    for i, item in enumerate(data):
        if not isinstance(item, dict):
            raise ValueError(f"Schema item #{i} must be an object")
        if "merged" not in item:
            raise ValueError(f"Schema item #{i} missing 'merged'")
        if not item.get("AS") and not item.get("InfraSluzba"):
            raise ValueError(f"Schema item #{i} must have at least one of 'AS' or 'InfraSluzba'")
        out.append(item)  # type: ignore[arg-type]
    return out


def _is_valid(meta: Optional[Meta]) -> bool:
    return bool(meta) and (meta.get("valid") is True)


def merge_attribute_metadata(
    as_meta: MetaMap,
    infra_meta: MetaMap,
    schema: List[SchemaItem],
    *,
    pass_through_unmapped: bool = True,
) -> MetaMap:
    out: MetaMap = {}
    mapped_as: Set[str] = set()
    mapped_infra: Set[str] = set()

    for rule in schema:
        as_key = rule.get("AS")
        infra_key = rule.get("InfraSluzba")
        merged_key = rule["merged"]

        if as_key:
            mapped_as.add(as_key)
        if infra_key:
            mapped_infra.add(infra_key)

        a = as_meta.get(as_key) if as_key else None
        b = infra_meta.get(infra_key) if infra_key else None

        a_ok = _is_valid(a)
        b_ok = _is_valid(b)
        if not (a_ok or b_ok):
            continue

        name = (
            rule.get("name")
            or (a.get("name") if a_ok else None)
            or (b.get("name") if b_ok else None)
            or merged_key
        )
        out[merged_key] = {"name": name, "valid": True}

    if not pass_through_unmapped:
        return out

    for k, m in as_meta.items():
        if k in mapped_as:
            continue
        if not _is_valid(m):
            continue
        out.setdefault(k, {"name": m.get("name", k), "valid": True})

    for k, m in infra_meta.items():
        if k in mapped_infra:
            continue
        if not _is_valid(m):
            continue
        out.setdefault(k, {"name": m.get("name", k), "valid": True})

    return out


def build_rename_maps(schema: List[SchemaItem]) -> Tuple[Dict[str, str], Dict[str, str]]:
    as_map: Dict[str, str] = {}
    is_map: Dict[str, str] = {}
    for rule in schema:
        mk = rule["merged"]
        ak = rule.get("AS")
        ik = rule.get("InfraSluzba")
        if ak:
            as_map[ak] = mk
        if ik:
            is_map[ik] = mk
    return as_map, is_map


def remap_attributes(
    attrs: AttrMap,
    rename_map: Dict[str, str],
    *,
    allowed_keys: Set[str] | None = None,
) -> AttrMap:
    out: AttrMap = {}

    # mapped keys
    for src, dst in rename_map.items():
        if src in attrs:
            out[dst] = attrs[src]

    # pass-through other keys
    mapped_src_keys = set(rename_map.keys())
    mapped_dst_keys = set(rename_map.values())

    for k, v in attrs.items():
        if k in mapped_src_keys:
            continue
        if k in mapped_dst_keys:
            out.setdefault(k, v)
        else:
            out.setdefault(k, v)

    # allowlist filtering
    if allowed_keys is not None:
        out = {k: v for k, v in out.items() if k in allowed_keys}

    return out


def remap_entities(
    entities: List[Dict[str, Any]],
    *,
    entity_kind: str,  # "AS" or "InfraSluzba"
    schema: List[SchemaItem],
    merged_meta: Dict[str, Any],
) -> List[Dict[str, Any]]:
    as_map, is_map = build_rename_maps(schema)
    rename_map = as_map if entity_kind == "AS" else is_map
    allowed_keys = set(merged_meta.keys())

    out: List[Dict[str, Any]] = []
    for e in entities:
        ne = dict(e)
        ne["attributes"] = remap_attributes(ne.get("attributes", {}) or {}, rename_map, allowed_keys=allowed_keys)
        out.append(ne)
    return out


# ----------------------------
# Table/Excel helpers
# ----------------------------

def _stringify_cell_value(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, (str, int, float, bool)):
        return str(x)
    if isinstance(x, list):
        return "; ".join(_stringify_cell_value(v) for v in x)
    if isinstance(x, dict):
        return json.dumps(x, ensure_ascii=False)
    return str(x)


def _build_attr_order(meta: Dict[str, Any], attr_order: List[str] | None) -> List[str]:
    all_keys = sorted(meta.keys())
    if not attr_order:
        return all_keys

    seen: Set[str] = set()
    out: List[str] = []

    for k in attr_order:
        if k in meta and k not in seen:
            out.append(k)
            seen.add(k)

    for k in all_keys:
        if k not in seen:
            out.append(k)

    return out


def _sort_key_for_attr(v: Any) -> str:
    return _stringify_cell_value(v).casefold()


def _unique_header_labels(meta: Dict[str, Any], order: List[str], incl_tech_name: bool = False) -> List[str]:
    # you chose always "name [key]" — keep that behavior
    labels: List[str] = []
    for k in order:
        base = meta[k].get("name", k)
        if incl_tech_name:
            labels.append(f"{base} [{k}]")
        else:
            labels.append(f"{base}")
    return labels


def _is_present_value(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip() != ""
    if isinstance(v, (list, tuple, set, dict)):
        return len(v) > 0
    return True


def write_excel(
    out_path: Path,
    data: List[Dict[str, Any]],
    meta: Dict[str, Any],
    *,
    attr_order: List[str] | None = None,
    sort_by: str | None = None,
    link_attr: str = "Gen_Profil_nazov",
    drop_param_threshold: float | None = 0,
    header_incl_tech_name: bool = False,
) -> None:
    order = _build_attr_order(meta, attr_order)

    # optionally drop rare/unused columns
    if drop_param_threshold is not None:
        total = len(data)
        counts: Dict[str, int] = {k: 0 for k in order}

        for thing in data:
            attrs = thing.get("attributes") or {}
            for k in order:
                if _is_present_value(attrs.get(k)):
                    counts[k] += 1

        if drop_param_threshold == 0:
            keep = {k for k, c in counts.items() if c > 0}
        else:
            thr = float(drop_param_threshold)
            if thr < 0:
                raise ValueError("drop_param_threshold must be >= 0, or None")
            keep = {k for k, c in counts.items() if (c / total) >= thr} if total > 0 else set()

        order = [k for k in order if k in keep]

    # optional sort
    if sort_by:
        data = sorted(data, key=lambda t: _sort_key_for_attr((t.get("attributes") or {}).get(sort_by)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Cloud services"

    header = ["Citype služby"] + _unique_header_labels(meta, order, incl_tech_name=header_incl_tech_name)
    ws.append(header)

    for col in range(1, len(header) + 1):
        c = ws.cell(row=1, column=col)
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"

    try:
        link_col = 2 + order.index(link_attr)
    except ValueError:
        link_col = None

    for thing in data:
        citype = thing.get("type", "")
        typ = "Aplikačná" if citype == "AS" else "Infraštruktúrna"
        attrs = thing.get("attributes") or {}

        row_values = [typ] + [_stringify_cell_value(attrs.get(k)) for k in order]
        ws.append(row_values)

        if link_col is not None:
            name_cell = ws.cell(row=ws.max_row, column=link_col)
            uuid = thing.get("uuid")
            if uuid and citype in ("AS", "InfraSluzba"):
                url = f"https://metais.slovensko.sk/ci/{citype}/{uuid}"
                name_cell.hyperlink = url
                name_cell.style = "Hyperlink"

    for col in range(1, len(header) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 22
        for r in range(2, ws.max_row + 1):
            ws.cell(r, col).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(out_path)


def standardize_data(
    data: List[Dict[str, Any]],
    meta: Dict[str, Any],
    *,
    attr_order: List[str] | None = None,
    sort_by: str | None = None,
    header_incl_tech_name: bool = False,
) -> Dict[str, Any]:
    order = _build_attr_order(meta, attr_order)

    if sort_by:
        data = sorted(data, key=lambda t: _sort_key_for_attr((t.get("attributes") or {}).get(sort_by)))

    header = ["Citype služby"] + _unique_header_labels(meta, order, incl_tech_name=header_incl_tech_name)

    rows: List[List[str]] = []
    for thing in data:
        citype = thing.get("type", "")
        typ = "Aplikačná" if citype == "AS" else "Infraštruktúrna"
        attrs = thing.get("attributes") or {}
        row = [typ] + [_stringify_cell_value(attrs.get(k)) for k in order]
        rows.append(row)

    return {"header": header, "rows": rows}
